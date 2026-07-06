from typing import Dict, Any
from fastapi import UploadFile
from openpyxl import load_workbook
from io import BytesIO
from sqlalchemy.orm import Session
from datetime import datetime
from models.historico import Ticket, MovimientoCaja


# ---------------------------------------------------------
# UTILIDADES
# ---------------------------------------------------------

def _to_float(v):
    """Convierte valores Excel a float de forma segura."""
    if v is None:
        return 0.0
    if isinstance(v, str):
        v = v.replace(",", ".").strip()
    try:
        return float(v)
    except:
        return 0.0


def _to_date(v):
    """Convierte fechas Excel a datetime."""
    if isinstance(v, datetime):
        return v
    if isinstance(v, float):  # número de serie Excel
        return datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(v) - 2)
    return None


# ---------------------------------------------------------
# LECTURA DEL EXCEL DETALLADO
# ---------------------------------------------------------

def _leer_hoja_principal(excel_file: UploadFile):
    contenido = excel_file.file.read()
    excel_file.file.seek(0)  # ← CRÍTICO: permite relecturas
    wb = load_workbook(BytesIO(contenido), data_only=True)
    ws = wb[wb.sheetnames[0]]
    return ws


def _mapear_cabeceras(ws) -> Dict[str, int]:
    cabeceras = {}
    for idx, cell in enumerate(ws[1], start=1):
        nombre = str(cell.value).strip().lower() if cell.value else ""
        nombre = nombre.replace("í", "i")  # ← normalización
        if nombre:
            cabeceras[nombre] = idx
    return cabeceras


# ---------------------------------------------------------
# IMPORTADOR DETALLADO (GUARDA EN BD)
# ---------------------------------------------------------

def importar_caja_excel(caja_excel: UploadFile, db: Session) -> Dict[str, Any]:
    try:
        ws = _leer_hoja_principal(caja_excel)
        cabeceras = _mapear_cabeceras(ws)

        print("CABECERAS DETECTADAS:", cabeceras)

        # Columnas necesarias (normalizadas)
        required = [
            "ticket", "fecha venta", "cliente", "empleado",
            "concepto", "categoria", "cantidad",
            "subtotal", "impuesto", "cuota", "descuento", "total"
        ]

        missing = [col for col in required if col not in cabeceras]
        if missing:
            return {
                "error": "Faltan columnas en el Excel",
                "faltan": missing,
                "cabeceras_detectadas": list(cabeceras.keys())
            }

        col_ticket       = cabeceras["ticket"]
        col_fecha_venta  = cabeceras["fecha venta"]
        col_cliente      = cabeceras["cliente"]
        col_empleado     = cabeceras["empleado"]
        col_concepto     = cabeceras["concepto"]
        col_categoria    = cabeceras["categoria"]
        col_cantidad     = cabeceras["cantidad"]
        col_subtotal     = cabeceras["subtotal"]
        col_impuesto     = cabeceras["impuesto"]
        col_cuota        = cabeceras["cuota"]
        col_descuento    = cabeceras["descuento"]
        col_total        = cabeceras["total"]
        col_forma_pago   = cabeceras.get("forma de pago")

        formas_pago_por_ticket: Dict[str, str] = {}
        tickets_cache: Dict[str, Ticket] = {}
        movimientos_count = 0

        # ---------------------------------------------------------
        # RECORRER TODAS LAS FILAS DEL EXCEL
        # ---------------------------------------------------------

        for row_idx in range(2, ws.max_row + 1):
            ticket_val = ws.cell(row=row_idx, column=col_ticket).value
            concepto = ws.cell(row=row_idx, column=col_concepto).value

            if ticket_val is None and concepto is None:
                continue

            ticket_codigo = str(ticket_val).strip() if ticket_val else ""

            # ---------------------------------------------------------
            # FILA "TOTAL TICKET"
            # ---------------------------------------------------------
            if concepto and str(concepto).strip().upper() == "TOTAL TICKET":
                forma_pago = ws.cell(row=row_idx, column=col_forma_pago).value if col_forma_pago else ""
                formas_pago_por_ticket[ticket_codigo] = str(forma_pago).strip() if forma_pago else ""

                total_ticket = _to_float(ws.cell(row=row_idx, column=col_total).value)
                fecha_venta = _to_date(ws.cell(row=row_idx, column=col_fecha_venta).value)
                cliente = ws.cell(row=row_idx, column=col_cliente).value
                empleado = ws.cell(row=row_idx, column=col_empleado).value

                movimientos_ticket = db.query(MovimientoCaja).join(Ticket).filter(
                    Ticket.ticket_codigo == ticket_codigo
                ).all()

                total_base = sum(m.subtotal or 0 for m in movimientos_ticket)
                total_iva = sum(m.iva_cuota or 0 for m in movimientos_ticket)

                ticket_obj = db.query(Ticket).filter(Ticket.ticket_codigo == ticket_codigo).first()
                if not ticket_obj:
                    ticket_obj = Ticket(
                        ticket_codigo=ticket_codigo,
                        fecha_venta=fecha_venta,
                        cliente=cliente,
                        empleado=empleado,
                        total_base=total_base,
                        total_iva=total_iva,
                        total_ticket=total_ticket,
                        forma_pago=formas_pago_por_ticket.get(ticket_codigo, "")
                    )
                    db.add(ticket_obj)
                else:
                    ticket_obj.fecha_venta = fecha_venta
                    ticket_obj.cliente = cliente
                    ticket_obj.empleado = empleado
                    ticket_obj.total_base = total_base
                    ticket_obj.total_iva = total_iva
                    ticket_obj.total_ticket = total_ticket
                    ticket_obj.forma_pago = formas_pago_por_ticket.get(ticket_codigo, "")

                db.commit()
                tickets_cache[ticket_codigo] = ticket_obj
                continue

            # ---------------------------------------------------------
            # FILA NORMAL → MOVIMIENTO
            # ---------------------------------------------------------

            fecha_venta = _to_date(ws.cell(row=row_idx, column=col_fecha_venta).value)
            cliente = ws.cell(row=row_idx, column=col_cliente).value
            empleado = ws.cell(row=row_idx, column=col_empleado).value
            categoria = ws.cell(row=row_idx, column=col_categoria).value
            cantidad = _to_float(ws.cell(row=row_idx, column=col_cantidad).value)
            subtotal = _to_float(ws.cell(row=row_idx, column=col_subtotal).value)
            impuesto = ws.cell(row=row_idx, column=col_impuesto).value
            cuota = _to_float(ws.cell(row=row_idx, column=col_cuota).value)
            descuento = _to_float(ws.cell(row=row_idx, column=col_descuento).value)
            total_linea = _to_float(ws.cell(row=row_idx, column=col_total).value)

            ticket_obj = tickets_cache.get(ticket_codigo)
            if not ticket_obj:
                ticket_obj = db.query(Ticket).filter(Ticket.ticket_codigo == ticket_codigo).first()
                if not ticket_obj:
                    ticket_obj = Ticket(
                        ticket_codigo=ticket_codigo,
                        fecha_venta=fecha_venta,
                        cliente=cliente,
                        empleado=empleado,
                        total_base=0.0,
                        total_iva=0.0,
                        total_ticket=0.0,
                        forma_pago=""
                    )
                    db.add(ticket_obj)
                    db.commit()
                tickets_cache[ticket_codigo] = ticket_obj

            movimiento = MovimientoCaja(
                ticket=ticket_obj,
                fecha_venta=fecha_venta,
                cliente=cliente,
                empleado=empleado,
                concepto=concepto,
                categoria=categoria,
                cantidad=cantidad,
                subtotal=subtotal,
                iva_porcentaje=str(impuesto or ""),
                iva_cuota=cuota,
                descuento=descuento,
                total_linea=total_linea,
            )

            db.add(movimiento)
            movimientos_count += 1

        db.commit()

        return {
            "tickets_creados": len(tickets_cache),
            "movimientos_historico": movimientos_count,
            "desconocidos": []
        }

    except Exception as e:
        return {
            "error": "Excepción interna en el importador",
            "detalle": str(e)
        }
